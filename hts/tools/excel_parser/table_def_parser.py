"""
テーブル定義書Excelファイルパーサー

openpyxlを使用してテーブル定義書をパースし、辞書形式およびMarkdown形式で出力する。
"""

from openpyxl import load_workbook
from typing import Dict, List, Any


def parse_table_definitions(filepath: str) -> Dict[str, Any]:
    """
    テーブル定義書Excelファイルをパースする

    Args:
        filepath: Excelファイルのパス

    Returns:
        {
            "overview": {...},  # テーブル一覧情報
            "tables": {
                "テーブル物理名": {
                    "logical_name": str,
                    "physical_name": str,
                    "created_date": str,
                    "modified_date": str,
                    "overview": str,
                    "columns": [
                        {
                            "no": int,
                            "logical_name": str,
                            "physical_name": str,
                            "data_type": str,
                            "length": str,
                            "not_null": bool,
                            "pk": bool,
                            "uk": bool,
                            "fk": bool,
                            "default": str,
                            "remarks": str
                        },
                        ...
                    ]
                },
                ...
            }
        }
    """
    wb = load_workbook(filepath, data_only=True)

    # 特殊シート（パース対象外）
    skip_sheets = {"表紙", "変更履歴", "テーブル一覧"}

    result = {
        "overview": _parse_table_list(wb),
        "tables": {}
    }

    # 各テーブル定義シートをパース
    for sheet_name in wb.sheetnames:
        if sheet_name not in skip_sheets:
            table_data = _parse_table_sheet(wb[sheet_name])
            if table_data:
                result["tables"][table_data["physical_name"]] = table_data

    return result


def _parse_table_list(wb) -> Dict[str, Any]:
    """
    テーブル一覧シートをパース

    Args:
        wb: openpyxl workbook

    Returns:
        テーブル一覧情報の辞書
    """
    if "テーブル一覧" not in wb.sheetnames:
        return {}

    sheet = wb["テーブル一覧"]
    tables = []

    # 行6以降がデータ行と仮定（ヘッダーは行5）
    row_idx = 6
    while True:
        no = sheet.cell(row=row_idx, column=2).value
        if no is None:
            break

        logical_name = sheet.cell(row=row_idx, column=5).value
        physical_name = sheet.cell(row=row_idx, column=16).value
        table_type = sheet.cell(row=row_idx, column=27).value
        overview = sheet.cell(row=row_idx, column=33).value

        tables.append({
            "no": no,
            "logical_name": logical_name or "",
            "physical_name": physical_name or "",
            "type": table_type or "",
            "overview": overview or ""
        })

        row_idx += 1

    return {"tables": tables}


def _parse_table_sheet(sheet) -> Dict[str, Any]:
    """
    個別テーブル定義シートをパース

    Args:
        sheet: openpyxl worksheet

    Returns:
        テーブル定義情報の辞書
    """
    # 行1-2からテーブルメタ情報を取得
    logical_name = sheet.cell(row=1, column=21).value or ""
    physical_name = sheet.cell(row=2, column=21).value or ""
    created_date = sheet.cell(row=1, column=35).value or ""
    modified_date = sheet.cell(row=1, column=49).value or ""

    # テーブル一覧から概要を取得（ここでは空文字）
    overview = ""

    # 行6以降のカラム定義をパース
    columns = []
    row_idx = 6

    while True:
        no = sheet.cell(row=row_idx, column=2).value
        if no is None:
            break

        col_logical_name = sheet.cell(row=row_idx, column=5).value or ""
        col_physical_name = sheet.cell(row=row_idx, column=14).value or ""
        data_type = sheet.cell(row=row_idx, column=23).value or ""
        length = sheet.cell(row=row_idx, column=29).value or ""
        not_null = sheet.cell(row=row_idx, column=32).value == "●"
        pk = sheet.cell(row=row_idx, column=35).value == "●"
        uk = sheet.cell(row=row_idx, column=37).value == "●"
        fk = sheet.cell(row=row_idx, column=39).value == "●"
        default = sheet.cell(row=row_idx, column=41).value or ""
        remarks = sheet.cell(row=row_idx, column=46).value or ""

        columns.append({
            "no": no,
            "logical_name": col_logical_name,
            "physical_name": col_physical_name,
            "data_type": data_type,
            "length": str(length) if length else "",
            "not_null": not_null,
            "pk": pk,
            "uk": uk,
            "fk": fk,
            "default": str(default) if default else "",
            "remarks": remarks
        })

        row_idx += 1

    return {
        "logical_name": logical_name,
        "physical_name": physical_name,
        "created_date": str(created_date) if created_date else "",
        "modified_date": str(modified_date) if modified_date else "",
        "overview": overview,
        "columns": columns
    }


def to_markdown(data: Dict[str, Any]) -> str:
    """
    パース結果をMarkdown形式に変換

    Args:
        data: parse_table_definitions()の戻り値

    Returns:
        Markdown形式の文字列
    """
    lines = []

    # テーブル一覧
    lines.append("# テーブル定義書\n")
    lines.append("## テーブル一覧\n")

    if data.get("overview", {}).get("tables"):
        lines.append("| No | 論理名 | 物理名 | 種別 | 概要 |")
        lines.append("|----|--------|--------|------|------|")

        for tbl in data["overview"]["tables"]:
            lines.append(
                f"| {tbl['no']} | {tbl['logical_name']} | {tbl['physical_name']} | "
                f"{tbl['type']} | {tbl['overview']} |"
            )

        lines.append("\n")

    # 各テーブル定義
    lines.append("## テーブル詳細\n")

    for table_name, table_info in data.get("tables", {}).items():
        lines.append(f"### {table_info['logical_name']} ({table_info['physical_name']})\n")
        lines.append(f"- **作成日**: {table_info['created_date']}")
        lines.append(f"- **変更日**: {table_info['modified_date']}")
        if table_info['overview']:
            lines.append(f"- **概要**: {table_info['overview']}")
        lines.append("")

        # カラム定義テーブル
        lines.append("| No | 論理名 | 物理名 | データ型 | 桁数 | NOT NULL | PK | UK | FK | デフォルト | 備考 |")
        lines.append("|----|--------|--------|----------|------|----------|----|----|----|-----------|----- |")

        for col in table_info["columns"]:
            nn = "●" if col["not_null"] else ""
            pk = "●" if col["pk"] else ""
            uk = "●" if col["uk"] else ""
            fk = "●" if col["fk"] else ""

            lines.append(
                f"| {col['no']} | {col['logical_name']} | {col['physical_name']} | "
                f"{col['data_type']} | {col['length']} | {nn} | {pk} | {uk} | {fk} | "
                f"{col['default']} | {col['remarks']} |"
            )

        lines.append("\n")

    return "\n".join(lines)

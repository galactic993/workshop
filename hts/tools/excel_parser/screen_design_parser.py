"""
画面設計書Excelパーサー
openpyxlを使用して画面設計書をパースし、構造化データとMarkdownに変換する
"""

from openpyxl import load_workbook
from typing import Dict, List, Any, Optional


def _get_cell_value(sheet, row: int, col: int) -> str:
    """セルの値を取得（Noneの場合は空文字を返す）"""
    cell = sheet.cell(row=row, column=col)
    value = cell.value
    return str(value).strip() if value is not None else ""


def _parse_metadata(sheet) -> Dict[str, str]:
    """共通ヘッダー情報を解析（行1-2）"""
    metadata = {}

    # 行1-2の各ラベル・値ペアを取得
    # col2=ラベル, col7=値, col16=ラベル, col21=値, col30=ラベル, col35=値, col44=ラベル, col49=値
    pairs = [
        (2, 7),   # 1列目
        (16, 21), # 2列目
        (30, 35), # 3列目
        (44, 49), # 4列目
    ]

    for row in [1, 2]:
        for label_col, value_col in pairs:
            label = _get_cell_value(sheet, row, label_col)
            value = _get_cell_value(sheet, row, value_col)
            if label:
                metadata[label] = value

    return metadata


def _parse_events(sheet) -> List[Dict[str, str]]:
    """イベント記述書を解析（行5ヘッダー、6以降データ）"""
    events = []

    # ヘッダー行確認（行5）
    event_no_header = _get_cell_value(sheet, 5, 2)
    if "EVENT" not in event_no_header.upper():
        return events

    # 行6以降のデータを読み取り
    row = 6
    while row <= sheet.max_row:
        event_no = _get_cell_value(sheet, row, 2)  # col2: EVENT No

        # EVENT Noが空の場合、次の行をチェック（空行スキップ）
        if not event_no:
            row += 1
            continue

        # EVENT で始まらない場合は終了
        if not event_no.startswith("EVENT"):
            break

        # イベントの処理内容を複数行から収集
        trigger = _get_cell_value(sheet, row, 7)
        type_value = _get_cell_value(sheet, row, 21)
        process_lines = []

        # 現在行と次の数行から処理内容を取得
        for offset in range(10):  # 最大10行先まで探索
            process_content = _get_cell_value(sheet, row + offset, 26)
            if process_content:
                process_lines.append(process_content)

            # 次のEVENTが来たら終了
            next_event = _get_cell_value(sheet, row + offset + 1, 2)
            if next_event and next_event.startswith("EVENT"):
                break

        event = {
            "event_no": event_no,
            "trigger": trigger,
            "type": type_value,
            "process_main": "\n".join(process_lines) if process_lines else "",
            "process_sub1": "",  # 仕様に従い空として扱う
            "process_sub2": "",  # 仕様に従い空として扱う
        }
        events.append(event)
        row += 1

    return events


def _parse_items(sheet) -> List[Dict[str, str]]:
    """項目記述書を解析（行4ヘッダー、5以降データ）"""
    items = []

    # ヘッダー行確認（行4）- "No"列を探す
    no_header = _get_cell_value(sheet, 4, 2)
    if not no_header or "No" not in no_header:
        return items

    # 行5以降のデータを読み取り
    row = 5
    empty_count = 0
    while row <= sheet.max_row:
        no = _get_cell_value(sheet, row, 2)  # col2: No

        # Noが空の場合はスキップ
        if not no:
            empty_count += 1
            row += 1
            # 連続20行空の場合は終了
            if empty_count >= 20:
                break
            continue

        # データが見つかったので empty_count をリセット
        empty_count = 0

        item = {
            "no": no,
            "name": _get_cell_value(sheet, row, 7),          # col7: 項目名
            "type": _get_cell_value(sheet, row, 18),         # col18: 種別
            "event_no": _get_cell_value(sheet, row, 25),     # col25: EVENT No
            "control_type": _get_cell_value(sheet, row, 30), # col30: 制御区分
            "control_detail": _get_cell_value(sheet, row, 35), # col35: 制御内容
            "message": _get_cell_value(sheet, row, 45),      # col45: メッセージ
        }
        items.append(item)
        row += 1

    return items


def _parse_db_values(sheet) -> List[Dict[str, str]]:
    """DB登録値シートを解析"""
    db_values = []

    # ヘッダー行を探す（"No."を含む行）
    header_row = None
    for row in range(1, min(20, sheet.max_row + 1)):
        cell_value = _get_cell_value(sheet, row, 2)
        if cell_value == "No." or cell_value == "No":
            header_row = row
            break

    if header_row is None:
        return db_values

    # ヘッダー行の次からデータを読み取り
    row = header_row + 1
    empty_count = 0
    while row <= sheet.max_row:
        no = _get_cell_value(sheet, row, 2)  # col2: No

        # Noが空の場合はスキップ
        if not no:
            empty_count += 1
            row += 1
            # 連続20行空の場合は終了
            if empty_count >= 20:
                break
            continue

        # データが見つかったので empty_count をリセット
        empty_count = 0

        db_value = {
            "no": no,
            "name": _get_cell_value(sheet, row, 7),          # col7: 項目名
            "field_name": _get_cell_value(sheet, row, 15),   # col15: フィールド名
            "condition": _get_cell_value(sheet, row, 25),    # col25: 条件
            "value": _get_cell_value(sheet, row, 40),        # col40: 登録値
        }
        db_values.append(db_value)
        row += 1

    return db_values


def _parse_messages(sheet) -> List[Dict[str, str]]:
    """メッセージ一覧シートを解析"""
    messages = []

    # ヘッダー行を探す（"No."を含む行）
    header_row = None
    for row in range(1, min(10, sheet.max_row + 1)):
        cell_value = _get_cell_value(sheet, row, 2)
        if cell_value == "No." or cell_value == "No":
            header_row = row
            break

    if header_row is None:
        return messages

    # ヘッダー行の次からデータを読み取り
    row = header_row + 1
    empty_count = 0
    while row <= sheet.max_row:
        no = _get_cell_value(sheet, row, 2)  # col2: No

        # Noが空の場合はスキップ
        if not no:
            empty_count += 1
            row += 1
            # 連続20行空の場合は終了
            if empty_count >= 20:
                break
            continue

        # データが見つかったので empty_count をリセット
        empty_count = 0

        message = {
            "no": no,
            "condition": _get_cell_value(sheet, row, 7),     # col7: 条件
            "type": _get_cell_value(sheet, row, 21),         # col21: 区分(エラー/サクセス)
            "message": _get_cell_value(sheet, row, 30),      # col30: メッセージ
        }
        messages.append(message)
        row += 1

    return messages


def parse_screen_design(filepath: str) -> Dict[str, Any]:
    """
    画面設計書Excelファイルをパースして構造化データを返す

    Args:
        filepath: Excelファイルのパス

    Returns:
        dict: メタデータ、イベント、項目、DB登録値、メッセージを含む辞書
    """
    wb = load_workbook(filepath, data_only=True)

    # メタデータは最初のシートから取得
    first_sheet = wb.worksheets[0]
    metadata = _parse_metadata(first_sheet)

    # 各シートから該当データを抽出
    events = []
    items = []
    db_values = []
    messages = []

    for sheet in wb.worksheets:
        sheet_title = sheet.title

        # イベント記述書シート
        if "イベント記述書" in sheet_title:
            events = _parse_events(sheet)

        # 項目記述書シート（完全一致で「項目記述書」のみ）
        elif sheet_title == "項目記述書":
            items = _parse_items(sheet)

        # DB登録値シート
        elif "DB登録値" in sheet_title or sheet_title == "DB登録値":
            db_values = _parse_db_values(sheet)

        # メッセージ一覧シート
        elif "メッセージ一覧" in sheet_title:
            messages = _parse_messages(sheet)

    result = {
        "metadata": metadata,
        "events": events,
        "items": items,
        "db_values": db_values,
        "messages": messages,
    }

    wb.close()
    return result


def to_markdown(data: Dict[str, Any]) -> str:
    """
    パース結果をMarkdown形式に変換

    Args:
        data: parse_screen_design()の戻り値

    Returns:
        str: Markdown形式のテキスト
    """
    lines = []

    # メタデータ
    lines.append("# 画面設計書")
    lines.append("")
    if data.get("metadata"):
        lines.append("## メタデータ")
        for key, value in data["metadata"].items():
            lines.append(f"- **{key}**: {value}")
        lines.append("")

    # イベント記述書
    if data.get("events"):
        lines.append("## イベント記述書")
        lines.append("")
        for event in data["events"]:
            lines.append(f"### {event['event_no']}")
            lines.append(f"- **トリガー**: {event['trigger']}")
            lines.append(f"- **区分**: {event['type']}")
            lines.append(f"- **処理メイン**: {event['process_main']}")
            if event['process_sub1']:
                lines.append(f"- **処理サブ1**: {event['process_sub1']}")
            if event['process_sub2']:
                lines.append(f"- **処理サブ2**: {event['process_sub2']}")
            lines.append("")

    # 項目記述書
    if data.get("items"):
        lines.append("## 項目記述書")
        lines.append("")
        lines.append("| No | 項目名 | 種別 | EVENT No | 制御区分 | 制御内容 | メッセージ |")
        lines.append("|----|--------|------|----------|----------|----------|------------|")
        for item in data["items"]:
            lines.append(
                f"| {item['no']} | {item['name']} | {item['type']} | "
                f"{item['event_no']} | {item['control_type']} | "
                f"{item['control_detail']} | {item['message']} |"
            )
        lines.append("")

    # DB登録値
    if data.get("db_values"):
        lines.append("## DB登録値")
        lines.append("")
        lines.append("| No | 項目名 | フィールド名 | 条件 | 登録値 |")
        lines.append("|----|--------|--------------|------|--------|")
        for db_val in data["db_values"]:
            lines.append(
                f"| {db_val['no']} | {db_val['name']} | {db_val['field_name']} | "
                f"{db_val['condition']} | {db_val['value']} |"
            )
        lines.append("")

    # メッセージ一覧
    if data.get("messages"):
        lines.append("## メッセージ一覧")
        lines.append("")
        lines.append("| No | 条件 | 区分 | メッセージ |")
        lines.append("|----|------|------|------------|")
        for msg in data["messages"]:
            lines.append(
                f"| {msg['no']} | {msg['condition']} | {msg['type']} | {msg['message']} |"
            )
        lines.append("")

    return "\n".join(lines)

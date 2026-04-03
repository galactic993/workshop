#!/usr/bin/env python3
"""
全設計書チェック一括実行スクリプト

~/.claude/skills/s1.5-*/scripts/check_*.py を自動検出して順次実行し、
統合レポートを出力する。

Usage:
    python run_all_checks.py <設計書ディレクトリ> [--output check_all_result.xlsx]
"""

import argparse
import re
import subprocess
import sys
import tempfile
import time
from pathlib import Path
from typing import Dict, List

import openpyxl
from openpyxl import load_workbook


# チェックスクリプトの実行順序（明示指定）
ORDERED_CHECKS = [
    's1.5-doc-format-pattern',
    's1.5-event-no',
    's1.5-message-format',
    's1.5-validation-consistency',
    's1.5-reference-spec',
    's1.5-screen-capture',
]

# 一括ランナー自身・補助スキル（scripts/check_*.py がない）
_SKIP_S1_5_GLOB = frozenset({
    's1.5-all-docs',
    's1.5-design-check-verify',
    's1.5-cowork-screen-capture-review',
})


def find_check_scripts(skills_dir: Path) -> List[Dict]:
    """
    s1.5-* スキルの check_*.py を検出する
    Returns: [{'skill': スキル名, 'script': スクリプトパス}, ...]
    """
    found = []

    # 順序付きで検索
    for skill_name in ORDERED_CHECKS:
        skill_dir = skills_dir / skill_name / 'scripts'
        if not skill_dir.is_dir():
            continue
        scripts = list(skill_dir.glob('check_*.py'))
        if scripts:
            found.append({'skill': skill_name, 'script': scripts[0]})

    # 順序リストにない s1.5-* スキルも追加（拡張性）
    for skill_dir in sorted(skills_dir.glob('s1.5-*')):
        skill_name = skill_dir.name
        if skill_name in _SKIP_S1_5_GLOB:
            continue
        if skill_name in ORDERED_CHECKS:
            continue
        scripts_dir = skill_dir / 'scripts'
        if not scripts_dir.is_dir():
            continue
        scripts = list(scripts_dir.glob('check_*.py'))
        if scripts:
            found.append({'skill': skill_name, 'script': scripts[0]})

    return found


def run_check(script_path: Path, design_dir: str, output_path: Path) -> Dict:
    """
    1つのチェックスクリプトを実行する
    Returns: {
        'skill': スキル名,
        'returncode': 終了コード,
        'stdout': 標準出力,
        'stderr': エラー出力,
        'duration': 実行時間(秒),
        'output_file': 出力Excelパス or None
    }
    """
    skill_name = script_path.parent.parent.name
    start = time.time()

    cmd = [
        sys.executable, str(script_path),
        design_dir,
        '--output', str(output_path)
    ]

    try:
        result = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
            timeout=300  # 5分タイムアウト
        )
        duration = time.time() - start

        return {
            'skill': skill_name,
            'script': str(script_path),
            'returncode': result.returncode,
            'stdout': result.stdout,
            'stderr': result.stderr,
            'duration': round(duration, 2),
            'output_file': output_path if output_path.exists() else None,
        }
    except subprocess.TimeoutExpired:
        return {
            'skill': skill_name,
            'script': str(script_path),
            'returncode': -1,
            'stdout': '',
            'stderr': 'タイムアウト（5分）',
            'duration': 300,
            'output_file': None,
        }
    except Exception as e:
        return {
            'skill': skill_name,
            'script': str(script_path),
            'returncode': -1,
            'stdout': '',
            'stderr': str(e),
            'duration': 0,
            'output_file': None,
        }


def parse_result_counts(stdout: str) -> Dict[str, int]:
    """標準出力からPASS/WARN/FAILの件数を解析する"""
    counts = {'PASS': 0, 'WARN': 0, 'FAIL': 0}
    for line in stdout.split('\n'):
        if line.startswith('結果:') or line.startswith('\n結果:'):
            # "結果: PASS=X, WARN=Y, FAIL=Z" 形式
            for key in counts:
                m = re.search(rf'{key}=(\d+)', line)
                if m:
                    counts[key] = int(m.group(1))
    return counts


def merge_excel_results(check_results: List[Dict], output_path: Path) -> None:
    """
    各チェックの結果ExcelをマージしてサマリーExcelを生成する
    """
    wb_out = openpyxl.Workbook()

    # サマリーシート
    ws_summary = wb_out.active
    ws_summary.title = 'サマリー'
    ws_summary.append([
        'チェック名', '実行結果', 'PASS件数', 'WARN件数', 'FAIL件数',
        '実行時間(秒)', '詳細'
    ])

    for cr in check_results:
        counts = parse_result_counts(cr['stdout'])
        if cr['returncode'] == 0:
            overall = 'PASS'
        elif cr['returncode'] == -1:
            overall = 'ERROR'
        else:
            overall = 'WARN/FAIL'

        detail = cr['stderr'] if cr['stderr'] else cr['stdout'].strip().split('\n')[-1] if cr['stdout'] else ''
        ws_summary.append([
            cr['skill'],
            overall,
            counts['PASS'],
            counts['WARN'],
            counts['FAIL'],
            cr['duration'],
            detail[:200] if detail else ''
        ])

    # 各チェックの詳細シートをマージ
    for cr in check_results:
        if cr['output_file'] is None:
            continue

        try:
            wb_src = load_workbook(cr['output_file'], data_only=True)
            for src_sheet in wb_src.worksheets:
                # シート名にスキル名プレフィックスを付与
                new_title = f"{cr['skill'][:15]}-{src_sheet.title}"[:31]

                ws_new = wb_out.create_sheet(title=new_title)
                for row in src_sheet.iter_rows(values_only=True):
                    ws_new.append([v if v is not None else '' for v in row])
            wb_src.close()
        except Exception as e:
            # エラーシートを追加
            ws_err = wb_out.create_sheet(title=f"{cr['skill'][:20]}-ERR"[:31])
            ws_err.append(['エラー', str(e)])

    wb_out.save(output_path)
    print(f'統合レポート出力: {output_path}')


def main():
    parser = argparse.ArgumentParser(description='全設計書チェック一括実行')
    parser.add_argument('design_dir', help='設計書ディレクトリのパス')
    parser.add_argument('--output', default='check_all_result.xlsx',
                        help='統合レポートファイルパス')
    parser.add_argument('--skills-dir', default=None,
                        help='スキルディレクトリのパス（デフォルト: ~/.claude/skills/）')
    args = parser.parse_args()

    design_dir = args.design_dir
    if not Path(design_dir).is_dir():
        print(f'エラー: ディレクトリが見つかりません: {design_dir}', file=sys.stderr)
        sys.exit(1)

    # スキルディレクトリを決定
    if args.skills_dir:
        skills_dir = Path(args.skills_dir)
    else:
        skills_dir = Path.home() / '.claude' / 'skills'

    if not skills_dir.is_dir():
        print(f'エラー: スキルディレクトリが見つかりません: {skills_dir}', file=sys.stderr)
        sys.exit(1)

    # チェックスクリプトを検出
    checks = find_check_scripts(skills_dir)
    if not checks:
        print('チェックスクリプトが見つかりませんでした。', file=sys.stderr)
        sys.exit(1)

    print(f'検出されたチェック: {len(checks)}件')
    for c in checks:
        print(f'  - {c["skill"]}: {c["script"]}')
    print()

    # 一時ディレクトリに各チェックの出力を保存
    check_results = []
    with tempfile.TemporaryDirectory() as tmpdir:
        for c in checks:
            skill_name = c['skill']
            script_path = c['script']
            output_file = Path(tmpdir) / f'{skill_name}_result.xlsx'

            print(f'実行中: {skill_name}')
            result = run_check(script_path, design_dir, output_file)
            check_results.append(result)

            status = 'OK' if result['returncode'] == 0 else 'NG'
            print(f'  [{status}] {skill_name} ({result["duration"]}秒)')
            if result['stderr']:
                print(f'  STDERR: {result["stderr"][:200]}')

        # 統合レポートを生成
        print('\n統合レポートを生成中...')
        merge_excel_results(check_results, Path(args.output))

    # サマリー表示
    print('\n=== 実行サマリー ===')
    total_pass = total_warn = total_fail = 0
    all_ok = True

    for cr in check_results:
        counts = parse_result_counts(cr['stdout'])
        total_pass += counts['PASS']
        total_warn += counts['WARN']
        total_fail += counts['FAIL']

        if cr['returncode'] != 0:
            all_ok = False
            status = 'FAIL'
        else:
            status = 'PASS'

        print(f'{cr["skill"]}: {status} (PASS={counts["PASS"]}, WARN={counts["WARN"]}, FAIL={counts["FAIL"]})')

    print(f'\n合計: PASS={total_pass}, WARN={total_warn}, FAIL={total_fail}')
    sys.exit(0 if all_ok else 1)


if __name__ == '__main__':
    main()

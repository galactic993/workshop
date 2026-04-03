#!/usr/bin/env python3
"""
メッセージフォーマットチェッカー

画面設計書のメッセージ一覧シートを対象に、
エラー/サクセスメッセージが定義済みフォーマットパターンに準拠しているかチェックする。

Usage:
    python check_message_format.py <設計書ディレクトリ> [--output result.xlsx]
"""

import argparse
import re
import sys
from pathlib import Path
from typing import Dict, List, Optional

import unicodedata

import openpyxl
from openpyxl import load_workbook


def norm(s): return unicodedata.normalize('NFC', str(s))


# メッセージフォーマットパターン定義
MESSAGE_PATTERNS = [
    {
        'id': 'MSG-AUTH',
        'category': 'エラー',
        'pattern': re.compile(r'.+権限がありません$'),
        'description': 'アクセス権限なし'
    },
    {
        'id': 'MSG-FETCH-FAIL',
        'category': 'エラー',
        'pattern': re.compile(r'.+の(取得|検索|読込)に失敗しました。時間を空けて再度お試しください$'),
        'description': '取得/検索失敗（再試行案内あり）'
    },
    {
        'id': 'MSG-PROC-FAIL',
        'category': 'エラー',
        'pattern': re.compile(r'.+に失敗しました$'),
        'description': '処理失敗'
    },
    {
        'id': 'MSG-INPUT-ERR',
        'category': 'エラー',
        'pattern': re.compile(r'^入力内容に誤りがあります。各項目をご確認ください$'),
        'description': '入力エラー（固定文言）'
    },
    {
        'id': 'MSG-NOT-FOUND',
        'category': 'サクセス',
        'pattern': re.compile(r'^該当する.+が見つかりません$'),
        'description': '該当なし'
    },
    {
        'id': 'MSG-SUCCESS',
        'category': 'サクセス',
        'pattern': re.compile(r'.+しました$'),
        'description': '処理成功'
    },
]


def get_cell_value(sheet, row: int, col: int) -> str:
    """セルの値を取得（Noneの場合は空文字）"""
    value = sheet.cell(row=row, column=col).value
    return str(value).strip() if value is not None else ""


def extract_messages(sheet) -> List[Dict]:
    """
    メッセージ一覧シートからメッセージを抽出する
    """
    messages = []

    # ヘッダー行を探す（"No."または"No"を含む行）
    header_row = None
    for row in range(1, min(15, sheet.max_row + 1)):
        val = get_cell_value(sheet, row, 2)
        if val in ('No.', 'No'):
            header_row = row
            break

    if header_row is None:
        return messages

    # ヘッダー行の次からデータを読み取り
    row = header_row + 1
    empty_count = 0
    while row <= sheet.max_row:
        no = get_cell_value(sheet, row, 2)
        if not no:
            empty_count += 1
            row += 1
            if empty_count >= 20:
                break
            continue
        empty_count = 0

        messages.append({
            'row': row,
            'no': no,
            'condition': get_cell_value(sheet, row, 7),   # col7: 条件
            'type': get_cell_value(sheet, row, 21),        # col21: 区分
            'message': get_cell_value(sheet, row, 30),     # col30: メッセージ
        })
        row += 1

    return messages


def classify_message(message_text: str) -> Optional[str]:
    """
    メッセージをパターンに照合して分類する
    Returns: パターンID or None（どのパターンにも合致しない場合）
    """
    for p in MESSAGE_PATTERNS:
        if p['pattern'].match(message_text):
            return p['id']
    return None


def check_file(filepath: Path) -> List[Dict]:
    """
    1つの画面設計書ファイルをチェックする
    """
    results = []
    filename = filepath.name

    try:
        wb = load_workbook(filepath, data_only=True)
    except Exception as e:
        return [{'ファイル名': filename, 'No': '-', '区分': '-', 'メッセージ': '-',
                 'パターンID': '-', '詳細': f'ファイル読み込みエラー: {e}', '判定': 'FAIL'}]

    found_msg_sheet = False
    for sheet in wb.worksheets:
        if 'メッセージ一覧' not in norm(sheet.title):
            continue
        found_msg_sheet = True
        messages = extract_messages(sheet)

        if not messages:
            results.append({
                'ファイル名': filename, 'No': '-', '区分': '-', 'メッセージ': '-',
                'パターンID': '-', '詳細': 'メッセージが存在しません', '判定': 'PASS'
            })
            continue

        for msg in messages:
            text = msg['message']
            if not text:
                continue

            pattern_id = classify_message(text)
            if pattern_id:
                results.append({
                    'ファイル名': filename,
                    'No': msg['no'],
                    '区分': msg['type'],
                    'メッセージ': text,
                    'パターンID': pattern_id,
                    '詳細': 'フォーマット準拠',
                    '判定': 'PASS'
                })
            else:
                results.append({
                    'ファイル名': filename,
                    'No': msg['no'],
                    '区分': msg['type'],
                    'メッセージ': text,
                    'パターンID': '要確認',
                    '詳細': 'いずれのフォーマットパターンにも合致しません',
                    '判定': 'WARN'
                })

    if not found_msg_sheet:
        results.append({
            'ファイル名': filename, 'No': '-', '区分': '-', 'メッセージ': '-',
            'パターンID': '-', '詳細': 'メッセージ一覧シートが見つかりません', '判定': 'WARN'
        })

    wb.close()
    return results


def write_report(all_results: List[Dict], output_path: Path) -> None:
    """Excelレポートを出力する"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'チェック結果'

    headers = ['ファイル名', 'No', '区分', 'メッセージ', 'パターンID', '詳細', '判定']
    ws.append(headers)

    for row in all_results:
        ws.append([row.get(h, '') for h in headers])

    # パターン一覧シート
    ws2 = wb.create_sheet('フォーマットパターン')
    ws2.append(['パターンID', '区分', '説明', '正規表現'])
    for p in MESSAGE_PATTERNS:
        ws2.append([p['id'], p['category'], p['description'], p['pattern'].pattern])

    wb.save(output_path)
    print(f'レポート出力: {output_path}')


def main():
    parser = argparse.ArgumentParser(description='メッセージフォーマットチェッカー')
    parser.add_argument('design_dir', help='設計書ディレクトリのパス')
    parser.add_argument('--output', default='check_message_format_result.xlsx',
                        help='出力レポートファイルパス')
    args = parser.parse_args()

    design_dir = Path(args.design_dir)
    if not design_dir.is_dir():
        print(f'エラー: ディレクトリが見つかりません: {design_dir}', file=sys.stderr)
        sys.exit(1)

    screen_files = [f for f in design_dir.glob('*.xlsx')
                    if '画面設計書' in norm(f.name) and not f.name.startswith('~$')]

    if not screen_files:
        print('画面設計書が見つかりませんでした。', file=sys.stderr)
        sys.exit(1)

    print(f'対象ファイル数: {len(screen_files)}')

    all_results = []
    has_failure = False

    for filepath in sorted(screen_files):
        print(f'チェック中: {filepath.name}')
        results = check_file(filepath)
        all_results.extend(results)
        for r in results:
            if r['判定'] in ('FAIL', 'WARN'):
                has_failure = True

    output_path = Path(args.output)
    write_report(all_results, output_path)

    pass_count = sum(1 for r in all_results if r['判定'] == 'PASS')
    warn_count = sum(1 for r in all_results if r['判定'] == 'WARN')
    fail_count = sum(1 for r in all_results if r['判定'] == 'FAIL')
    print(f'\n結果: PASS={pass_count}, WARN={warn_count}, FAIL={fail_count}')

    sys.exit(1 if has_failure else 0)


if __name__ == '__main__':
    main()

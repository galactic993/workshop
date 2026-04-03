#!/usr/bin/env python3
"""最小の画面設計書・テーブル定義書を生成し、s1.5-* チェックスクリプトが動作するか検証する。"""
from pathlib import Path

from openpyxl import Workbook


def build_table_def(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "顧客マスタ"
    ws.cell(1, 21, "顧客マスタ")
    ws.cell(2, 21, "M_CUSTOMER")
    ws.cell(6, 2, 1)
    ws.cell(6, 5, "顧客番号")
    ws.cell(7, 2, 2)
    ws.cell(7, 5, "氏名")
    wb.save(path)


def build_screen(path: Path) -> None:
    wb = Workbook()
    names = [
        "表紙", "変更履歴", "画面概要", "イベント記述書", "項目記述書",
        "参照仕様", "DB登録値", "表示切替項目", "メッセージ一覧",
    ]
    wb.remove(wb.active)
    for t in names:
        wb.create_sheet(t)

    ev = wb["イベント記述書"]
    ev.cell(6, 2, "EVENT0001")
    ev.cell(7, 2, "EVENT0002")

    items = wb["項目記述書"]
    items.cell(4, 2, "No.")
    items.cell(5, 2, 1)
    items.cell(5, 7, "顧客番号")
    items.cell(5, 25, "EVENT0001")
    items.cell(5, 35, "必須")

    ref = wb["参照仕様"]
    ref.cell(2, 2, "◆顧客マスタ")
    ref.cell(3, 2, "抽出テーブル")
    ref.cell(3, 7, "顧客マスタ")
    ref.cell(4, 2, "抽出項目")
    ref.cell(4, 7, "顧客番号")

    msg = wb["メッセージ一覧"]
    msg.cell(8, 2, "No.")
    msg.cell(9, 2, 1)
    msg.cell(9, 7, "必須エラー")
    msg.cell(9, 21, "エラー")
    msg.cell(9, 30, "入力内容に誤りがあります。各項目をご確認ください")
    msg.cell(10, 2, 2)
    msg.cell(10, 7, "完了")
    msg.cell(10, 21, "サクセス")
    msg.cell(10, 30, "登録しました")

    wb.save(path)


def main() -> None:
    root = Path(__file__).resolve().parent.parent / "fixtures"
    root.mkdir(parents=True, exist_ok=True)
    build_table_def(root / "テーブル定義書_共.xlsx")
    build_screen(root / "サンプル画面設計書.xlsx")
    print(f"出力: {root}")


if __name__ == "__main__":
    main()

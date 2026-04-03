#!/usr/bin/env python3
"""
参照仕様整合性チェッカー

画面設計書の参照仕様シートに記載されたテーブル名・カラム名が、
テーブル定義書に実在するか検証する。

Usage:
    python check_reference_spec.py <設計書ディレクトリ> [--output result.xlsx]
"""

import argparse
import sys
from pathlib import Path
from typing import Dict, List, Set, Tuple

import unicodedata

import openpyxl
from openpyxl import load_workbook


def norm(s): return unicodedata.normalize('NFC', str(s))


def get_cell_value(sheet, row: int, col: int) -> str:
    value = sheet.cell(row=row, column=col).value
    return str(value).strip() if value is not None else ""


def build_table_dict(table_def_files: List[Path]) -> Dict[str, Set[str]]:
    """
    テーブル定義書からテーブル論理名 -> カラム論理名セットの辞書を構築する
    Returns: {テーブル論理名: {カラム論理名, ...}, ...}
    """
    table_dict: Dict[str, Set[str]] = {}
    skip_sheets = {'表紙', '変更履歴', 'テーブル一覧'}

    for filepath in table_def_files:
        try:
            wb = load_workbook(filepath, data_only=True)
        except Exception as e:
            print(f'テーブル定義書読み込みエラー: {filepath.name}: {e}', file=sys.stderr)
            continue

        for sheet_name in wb.sheetnames:
            if sheet_name in skip_sheets:
                continue

            sheet = wb[sheet_name]
            # 行1: 論理名がcol21、行2: 物理名がcol21
            logical_name = get_cell_value(sheet, 1, 21)
            if not logical_name:
                logical_name = sheet_name

            columns: Set[str] = set()
            # 行6以降: カラム定義（col5: カラム論理名）
            for row in range(6, sheet.max_row + 1):
                no = sheet.cell(row=row, column=2).value
                if no is None:
                    break
                col_logical = get_cell_value(sheet, row, 5)
                if col_logical:
                    columns.add(col_logical)

            if logical_name:
                table_dict[logical_name] = columns

        wb.close()

    return table_dict


def extract_reference_specs(sheet) -> List[Dict]:
    """
    参照仕様シートから参照テーブル・カラム情報を抽出する
    ◆から始まる行をセクション区切りとして解析
    """
    refs = []
    current_section = None

    for row in range(1, sheet.max_row + 1):
        # col2から読み取り
        val = get_cell_value(sheet, row, 2)
        if not val:
            continue

        # ◆セクション開始（テーブル名に相当）
        if val.startswith('◆'):
            current_section = val.lstrip('◆').strip()
            continue

        # 「抽出テーブル」「参照テーブル」行を検出
        if any(kw in val for kw in ['抽出テーブル', '参照テーブル', 'テーブル名']):
            table_name = get_cell_value(sheet, row, 7)  # col7に値
            if table_name and current_section:
                refs.append({
                    'row': row,
                    'section': current_section,
                    'type': 'テーブル',
                    'value': table_name,
                })
            continue

        # 「抽出項目」「カラム」行を検出
        if any(kw in val for kw in ['抽出項目', 'カラム', '項目名']):
            col_name = get_cell_value(sheet, row, 7)  # col7に値
            if col_name and current_section:
                refs.append({
                    'row': row,
                    'section': current_section,
                    'type': 'カラム',
                    'value': col_name,
                })
            continue

    return refs


def check_file(filepath: Path, table_dict: Dict[str, Set[str]]) -> List[Dict]:
    """1つの画面設計書ファイルをチェックする"""
    results = []
    filename = filepath.name

    try:
        wb = load_workbook(filepath, data_only=True)
    except Exception as e:
        return [{'ファイル名': filename, 'セクション': '-', '種別': '-',
                 '参照値': '-', '行番号': '-', '詳細': f'読み込みエラー: {e}', '判定': 'FAIL'}]

    found_ref_sheet = False
    for sheet in wb.worksheets:
        if '参照仕様' not in norm(sheet.title):
            continue
        found_ref_sheet = True
        refs = extract_reference_specs(sheet)

        # テーブル名と対応するカラムを関連付けてチェック
        current_table_logical = None
        for ref in refs:
            if ref['type'] == 'テーブル':
                # テーブル名を論理名辞書で検索
                found_table = ref['value'] in table_dict
                if found_table:
                    current_table_logical = ref['value']
                    results.append({
                        'ファイル名': filename,
                        'セクション': ref['section'],
                        '種別': 'テーブル',
                        '参照値': ref['value'],
                        '行番号': ref['row'],
                        '詳細': 'テーブル定義書に存在します',
                        '判定': 'PASS'
                    })
                else:
                    current_table_logical = None
                    # 部分一致で探す
                    partial_matches = [t for t in table_dict.keys() if ref['value'] in t or t in ref['value']]
                    detail = f"テーブル定義書に見つかりません"
                    if partial_matches:
                        detail += f"（候補: {', '.join(partial_matches[:3])}）"
                    results.append({
                        'ファイル名': filename,
                        'セクション': ref['section'],
                        '種別': 'テーブル',
                        '参照値': ref['value'],
                        '行番号': ref['row'],
                        '詳細': detail,
                        '判定': 'WARN'
                    })

            elif ref['type'] == 'カラム' and current_table_logical:
                # カラム名チェック
                columns = table_dict.get(current_table_logical, set())
                found_col = ref['value'] in columns
                if found_col:
                    results.append({
                        'ファイル名': filename,
                        'セクション': ref['section'],
                        '種別': 'カラム',
                        '参照値': ref['value'],
                        '行番号': ref['row'],
                        '詳細': f'{current_table_logical}のカラムとして存在します',
                        '判定': 'PASS'
                    })
                else:
                    # 部分一致で探す
                    partial = [c for c in columns if ref['value'] in c or c in ref['value']]
                    detail = f"{current_table_logical}にカラムが見つかりません"
                    if partial:
                        detail += f"（候補: {', '.join(list(partial)[:3])}）"
                    results.append({
                        'ファイル名': filename,
                        'セクション': ref['section'],
                        '種別': 'カラム',
                        '参照値': ref['value'],
                        '行番号': ref['row'],
                        '詳細': detail,
                        '判定': 'WARN'
                    })

    if not found_ref_sheet:
        results.append({
            'ファイル名': filename, 'セクション': '-', '種別': '-',
            '参照値': '-', '行番号': '-', '詳細': '参照仕様シートが見つかりません', '判定': 'WARN'
        })

    if not results:
        results.append({
            'ファイル名': filename, 'セクション': '-', '種別': '-',
            '参照値': '-', '行番号': '-', '詳細': '参照仕様の参照エントリが見つかりません', '判定': 'PASS'
        })

    wb.close()
    return results


def write_report(all_results: List[Dict], output_path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'チェック結果'

    headers = ['ファイル名', 'セクション', '種別', '参照値', '行番号', '詳細', '判定']
    ws.append(headers)
    for row in all_results:
        ws.append([row.get(h, '') for h in headers])

    wb.save(output_path)
    print(f'レポート出力: {output_path}')


def main():
    parser = argparse.ArgumentParser(description='参照仕様整合性チェッカー')
    parser.add_argument('design_dir', help='設計書ディレクトリのパス')
    parser.add_argument('--output', default='check_reference_spec_result.xlsx')
    args = parser.parse_args()

    design_dir = Path(args.design_dir)
    if not design_dir.is_dir():
        print(f'エラー: {design_dir}', file=sys.stderr)
        sys.exit(1)

    # テーブル定義書を収集（共・売・編）
    table_def_files = [f for f in design_dir.glob('*.xlsx')
                       if 'テーブル定義書' in norm(f.name) and not f.name.startswith('~$')]

    if not table_def_files:
        print('テーブル定義書が見つかりませんでした。', file=sys.stderr)
        sys.exit(1)

    print(f'テーブル定義書: {[f.name for f in table_def_files]}')

    # テーブル辞書構築
    table_dict = build_table_dict(table_def_files)
    print(f'テーブル定義数: {len(table_dict)}')

    # 画面設計書を収集
    screen_files = [f for f in design_dir.glob('*.xlsx')
                    if '画面設計書' in norm(f.name) and not f.name.startswith('~$')]

    if not screen_files:
        print('画面設計書が見つかりませんでした。', file=sys.stderr)
        sys.exit(1)

    print(f'対象ファイル数: {len(screen_files)}')

    all_results = []
    has_failure = False

    for filepath in sorted(screen_files):
        print(f'チェック中: {filepath.name}')
        results = check_file(filepath, table_dict)
        all_results.extend(results)
        for r in results:
            if r['判定'] in ('FAIL', 'WARN'):
                has_failure = True

    write_report(all_results, Path(args.output))

    pass_count = sum(1 for r in all_results if r['判定'] == 'PASS')
    warn_count = sum(1 for r in all_results if r['判定'] == 'WARN')
    fail_count = sum(1 for r in all_results if r['判定'] == 'FAIL')
    print(f'\n結果: PASS={pass_count}, WARN={warn_count}, FAIL={fail_count}')

    sys.exit(1 if has_failure else 0)


if __name__ == '__main__':
    main()

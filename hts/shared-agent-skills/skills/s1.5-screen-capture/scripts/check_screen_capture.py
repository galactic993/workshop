#!/usr/bin/env python3
"""
画面キャプチャ突合チェッカー

画面設計書の画面概要シートに埋め込まれたキャプチャ画像内の項目名・ボタンラベルと、
項目記述書の項目名リストを突合する。

Usage:
    python check_screen_capture.py <設計書ディレクトリ> [--output result.xlsx]
"""

import argparse
import base64
import json
import subprocess
import sys
import tempfile
from pathlib import Path
from typing import Dict, List, Optional

import unicodedata

import openpyxl
from openpyxl import load_workbook


def norm(s): return unicodedata.normalize('NFC', str(s))


def get_cell_value(sheet, row: int, col: int) -> str:
    value = sheet.cell(row=row, column=col).value
    return str(value).strip() if value is not None else ""


def check_claude_cli() -> bool:
    """claude CLIが利用可能か確認する"""
    try:
        result = subprocess.run(['claude', '--version'], capture_output=True, timeout=5)
        return result.returncode == 0
    except (subprocess.TimeoutExpired, FileNotFoundError):
        return False


def extract_images_from_sheet(sheet) -> List[bytes]:
    """
    シートから埋め込み画像を抽出する
    Returns: 画像バイナリのリスト
    """
    images = []
    try:
        for img in getattr(sheet, '_images', []):
            if hasattr(img, '_data'):
                data = img._data()
                if data:
                    images.append(data)
            elif hasattr(img, 'ref'):
                pass
    except Exception:
        pass
    return images


def read_items_with_claude(image_data: bytes) -> List[str]:
    """
    Claude CLIを使って画像から項目名・ボタンラベルを読み取る
    """
    try:
        with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as f:
            f.write(image_data)
            tmp_path = f.name

        prompt = (
            "この画面キャプチャ画像から、フォームの項目名（ラベル）とボタンラベルを全て抽出してください。"
            "JSONの配列形式（[\"項目名1\", \"項目名2\", ...]）で出力してください。"
            "説明文は不要です。配列のみ出力してください。"
        )

        result = subprocess.run(
            ['claude', '-p', prompt, '--image', tmp_path],
            capture_output=True, text=True, timeout=60
        )

        Path(tmp_path).unlink(missing_ok=True)

        if result.returncode != 0:
            return []

        output = result.stdout.strip()
        # JSON配列を抽出
        start = output.find('[')
        end = output.rfind(']')
        if start != -1 and end != -1:
            arr = json.loads(output[start:end+1])
            return [str(x) for x in arr]

    except Exception:
        pass
    return []


def extract_item_names_from_items_sheet(sheet) -> List[str]:
    """項目記述書から項目名リストを抽出する"""
    items = []

    no_header = get_cell_value(sheet, 4, 2)
    if not no_header or 'No' not in no_header:
        return items

    row = 5
    empty_count = 0
    while row <= sheet.max_row:
        no = get_cell_value(sheet, row, 2)
        if not no:
            empty_count += 1
            row += 1
            if empty_count >= 20:
                break
            continue
        empty_count = 0

        name = get_cell_value(sheet, row, 7)  # col7: 項目名
        if name:
            items.append(name)
        row += 1

    return items


def check_file(filepath: Path, claude_available: bool) -> List[Dict]:
    """1つの画面設計書ファイルをチェックする"""
    results = []
    filename = filepath.name

    try:
        wb = load_workbook(filepath, data_only=True)
    except Exception as e:
        return [{'ファイル名': filename, '種別': 'エラー', '項目名': '-',
                 '詳細': f'読み込みエラー: {e}', '判定': 'FAIL'}]

    # 項目記述書から項目名を収集
    item_names = set()
    for sheet in wb.worksheets:
        if norm(sheet.title) == '項目記述書':
            item_names.update(extract_item_names_from_items_sheet(sheet))

    if not item_names:
        wb.close()
        return [{'ファイル名': filename, '種別': 'INFO', '項目名': '-',
                 '詳細': '項目記述書に項目名が見つかりません', '判定': 'WARN'}]

    # 画面概要シートから画像を抽出
    capture_items = []
    has_overview_sheet = False
    for sheet in wb.worksheets:
        if '画面概要' not in norm(sheet.title):
            continue
        has_overview_sheet = True

        if not claude_available:
            results.append({
                'ファイル名': filename, '種別': 'SKIP', '項目名': '-',
                '詳細': 'claude CLIが利用できないため画像読み取りをスキップしました',
                '判定': 'WARN'
            })
            break

        images = extract_images_from_sheet(sheet)
        if not images:
            results.append({
                'ファイル名': filename, '種別': 'INFO', '項目名': '-',
                '詳細': '画面概要シートに埋め込み画像が見つかりません',
                '判定': 'WARN'
            })
            break

        for i, img_data in enumerate(images):
            print(f'  画像{i+1}を解析中...')
            read_items = read_items_with_claude(img_data)
            capture_items.extend(read_items)

    wb.close()

    if not has_overview_sheet:
        results.append({
            'ファイル名': filename, '種別': 'INFO', '項目名': '-',
            '詳細': '画面概要シートが見つかりません', '判定': 'WARN'
        })
        return results

    if not capture_items or not claude_available:
        return results

    # 突合チェック
    capture_set = set(capture_items)

    # 画像にあるが項目記述書にない項目
    only_in_capture = capture_set - item_names
    for item in sorted(only_in_capture):
        results.append({
            'ファイル名': filename, '種別': '画像のみ', '項目名': item,
            '詳細': '画面キャプチャにあるが項目記述書に見つかりません',
            '判定': 'WARN'
        })

    # 項目記述書にあるが画像にない項目
    only_in_items = item_names - capture_set
    for item in sorted(only_in_items):
        results.append({
            'ファイル名': filename, '種別': '記述書のみ', '項目名': item,
            '詳細': '項目記述書にあるが画面キャプチャで確認できません',
            '判定': 'WARN'
        })

    # 両方に存在する項目
    common = capture_set & item_names
    for item in sorted(common):
        results.append({
            'ファイル名': filename, '種別': '一致', '項目名': item,
            '詳細': '画面キャプチャと項目記述書の両方に存在します',
            '判定': 'PASS'
        })

    if not results:
        results.append({
            'ファイル名': filename, '種別': '突合完了', '項目名': '-',
            '詳細': f'項目記述書: {len(item_names)}件、画面キャプチャ: {len(capture_items)}件',
            '判定': 'PASS'
        })

    return results


def write_report(all_results: List[Dict], output_path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'チェック結果'

    headers = ['ファイル名', '種別', '項目名', '詳細', '判定']
    ws.append(headers)
    for row in all_results:
        ws.append([row.get(h, '') for h in headers])

    wb.save(output_path)
    print(f'レポート出力: {output_path}')


def main():
    parser = argparse.ArgumentParser(description='画面キャプチャ突合チェッカー')
    parser.add_argument('design_dir', help='設計書ディレクトリのパス')
    parser.add_argument('--output', default='check_screen_capture_result.xlsx')
    args = parser.parse_args()

    design_dir = Path(args.design_dir)
    if not design_dir.is_dir():
        print(f'エラー: {design_dir}', file=sys.stderr)
        sys.exit(1)

    claude_available = check_claude_cli()
    if not claude_available:
        print('警告: claude CLIが利用できません。画像読み取りをスキップします。')

    screen_files = [f for f in design_dir.glob('*.xlsx')
                    if '画面設計書' in norm(f.name) and not f.name.startswith('~$')]

    if not screen_files:
        print('画面設計書が見つかりませんでした。', file=sys.stderr)
        sys.exit(1)

    print(f'対象ファイル数: {len(screen_files)}')

    all_results = []
    has_failure = False

    for filepath in sorted(screen_files):
        print(f'チェック中: {filepath.name}')
        results = check_file(filepath, claude_available)
        all_results.extend(results)
        for r in results:
            if r['判定'] in ('FAIL', 'WARN'):
                has_failure = True

    write_report(all_results, Path(args.output))

    pass_count = sum(1 for r in all_results if r['判定'] == 'PASS')
    warn_count = sum(1 for r in all_results if r['判定'] == 'WARN')
    fail_count = sum(1 for r in all_results if r['判定'] == 'FAIL')
    print(f'\n結果: PASS={pass_count}, WARN={warn_count}, FAIL={fail_count}')

    sys.exit(1 if has_failure else 0)


if __name__ == '__main__':
    main()

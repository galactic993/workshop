#!/usr/bin/env python3
"""
EVENT No採番整合性チェッカー

画面設計書のイベント記述書・項目記述書シートを対象に、
EVENT No（EVENT0001形式）の採番整合性をチェックする。

Usage:
    python check_event_no.py <設計書ディレクトリ> [--output result.xlsx]
"""

import argparse
import re
import sys
from pathlib import Path
from typing import Dict, List, Tuple

import unicodedata

import openpyxl
from openpyxl import load_workbook


def norm(s): return unicodedata.normalize('NFC', str(s))


def get_cell_value(sheet, row: int, col: int) -> str:
    """セルの値を取得（Noneの場合は空文字）"""
    value = sheet.cell(row=row, column=col).value
    return str(value).strip() if value is not None else ""


def extract_event_nos_from_event_sheet(sheet) -> List[Tuple[str, int]]:
    """
    イベント記述書シートからEVENT Noと行番号を抽出する
    Returns: [(event_no, row_number), ...]
    """
    result = []
    pattern = re.compile(r'^EVENT\d{4}$')

    for row in range(6, sheet.max_row + 1):
        val = get_cell_value(sheet, row, 2)
        if pattern.match(val):
            result.append((val, row))

    return result


def extract_event_nos_from_item_sheet(sheet) -> List[Tuple[str, int]]:
    """
    項目記述書シートからEVENT Noと行番号を抽出する（col25）
    Returns: [(event_no, row_number), ...]
    """
    result = []
    pattern = re.compile(r'EVENT\d{4}')

    for row in range(5, sheet.max_row + 1):
        val = get_cell_value(sheet, row, 25)
        if not val:
            continue
        # カンマ区切り等で複数EVENT Noがある場合も対応
        matches = pattern.findall(val)
        for m in matches:
            result.append((m, row))

    return result


def check_event_nos(event_nos: List[Tuple[str, int]]) -> List[Dict]:
    """
    EVENT No一覧の欠番・重複をチェックする
    Returns: チェック結果のリスト
    """
    issues = []

    if not event_nos:
        return issues

    # 番号を抽出してソート
    nums = []
    seen = {}
    for event_no, row in event_nos:
        num = int(event_no[5:])
        if event_no in seen:
            issues.append({
                'check_type': '重複',
                'event_no': event_no,
                'row': row,
                'detail': f'{event_no} が重複しています（初出行: {seen[event_no]}）',
                'judgment': 'FAIL'
            })
        else:
            seen[event_no] = row
            nums.append(num)

    if not nums:
        return issues

    nums_sorted = sorted(nums)
    min_num = nums_sorted[0]
    max_num = nums_sorted[-1]

    # 欠番チェック
    existing = set(nums_sorted)
    for n in range(min_num, max_num + 1):
        if n not in existing:
            event_no = f'EVENT{n:04d}'
            issues.append({
                'check_type': '欠番',
                'event_no': event_no,
                'row': '-',
                'detail': f'{event_no} が欠番です',
                'judgment': 'WARN'
            })

    return issues


def generate_renumbering_plan(event_nos: List[Tuple[str, int]]) -> List[Dict]:
    """
    連番で再採番した場合の新旧対応表を生成する
    """
    plan = []
    sorted_nos = sorted(set(en for en, _ in event_nos))

    for i, old_no in enumerate(sorted_nos, start=1):
        new_no = f'EVENT{i:04d}'
        if old_no != new_no:
            plan.append({'旧EVENT No': old_no, '新EVENT No': new_no})

    return plan


def check_file(filepath: Path) -> List[Dict]:
    """
    1つの画面設計書ファイルをチェックする
    """
    results = []
    filename = filepath.name

    try:
        wb = load_workbook(filepath, data_only=True)
    except Exception as e:
        return [{'ファイル名': filename, 'チェック種別': 'エラー', 'EVENT No': '-',
                 '行番号': '-', '詳細': f'ファイル読み込みエラー: {e}', '判定': 'FAIL'}]

    event_nos_in_event_sheet = []
    event_nos_in_item_sheet = []

    for sheet in wb.worksheets:
        title = sheet.title
        if 'イベント記述書' in norm(title):
            event_nos_in_event_sheet.extend(extract_event_nos_from_event_sheet(sheet))
        elif norm(title) == '項目記述書':
            event_nos_in_item_sheet.extend(extract_event_nos_from_item_sheet(sheet))

    wb.close()

    if not event_nos_in_event_sheet:
        results.append({
            'ファイル名': filename, 'チェック種別': 'INFO',
            'EVENT No': '-', '行番号': '-',
            '詳細': 'イベント記述書シートが見つからないかEVENT Noが存在しません',
            '判定': 'WARN'
        })
        return results

    # 欠番・重複チェック
    issues = check_event_nos(event_nos_in_event_sheet)
    for issue in issues:
        results.append({
            'ファイル名': filename,
            'チェック種別': issue['check_type'],
            'EVENT No': issue['event_no'],
            '行番号': issue['row'],
            '詳細': issue['detail'],
            '判定': issue['judgment']
        })

    # 項目記述書からの未定義参照チェック
    defined_set = set(en for en, _ in event_nos_in_event_sheet)
    for ref_no, row in event_nos_in_item_sheet:
        if ref_no not in defined_set:
            results.append({
                'ファイル名': filename,
                'チェック種別': '未定義参照',
                'EVENT No': ref_no,
                '行番号': row,
                '詳細': f'項目記述書の{row}行目で参照されているが、イベント記述書に定義がありません',
                '判定': 'FAIL'
            })

    if not results:
        results.append({
            'ファイル名': filename,
            'チェック種別': '採番チェック',
            'EVENT No': '-',
            '行番号': '-',
            '詳細': f'EVENT No: {len(event_nos_in_event_sheet)}件すべて正常です',
            '判定': 'PASS'
        })

    return results


def write_report(all_results: List[Dict], renumbering_plans: Dict[str, List[Dict]],
                 output_path: Path) -> None:
    """Excelレポートを出力する"""
    wb = openpyxl.Workbook()

    # メインシート
    ws = wb.active
    ws.title = 'チェック結果'

    headers = ['ファイル名', 'チェック種別', 'EVENT No', '行番号', '詳細', '判定']
    ws.append(headers)

    for row in all_results:
        ws.append([row.get(h, '') for h in headers])

    # 振り直し案シート
    if renumbering_plans:
        ws2 = wb.create_sheet('振り直し案')
        ws2.append(['ファイル名', '旧EVENT No', '新EVENT No'])
        for filename, plan in renumbering_plans.items():
            for p in plan:
                ws2.append([filename, p['旧EVENT No'], p['新EVENT No']])

    wb.save(output_path)
    print(f'レポート出力: {output_path}')


def main():
    parser = argparse.ArgumentParser(description='EVENT No採番整合性チェッカー')
    parser.add_argument('design_dir', help='設計書ディレクトリのパス')
    parser.add_argument('--output', default='check_event_no_result.xlsx',
                        help='出力レポートファイルパス')
    args = parser.parse_args()

    design_dir = Path(args.design_dir)
    if not design_dir.is_dir():
        print(f'エラー: ディレクトリが見つかりません: {design_dir}', file=sys.stderr)
        sys.exit(1)

    # 画面設計書を検索
    screen_files = [f for f in design_dir.glob('*.xlsx')
                    if '画面設計書' in norm(f.name) and not f.name.startswith('~$')]

    if not screen_files:
        print('画面設計書が見つかりませんでした。', file=sys.stderr)
        sys.exit(1)

    print(f'対象ファイル数: {len(screen_files)}')

    all_results = []
    renumbering_plans = {}
    has_failure = False

    for filepath in sorted(screen_files):
        print(f'チェック中: {filepath.name}')
        results = check_file(filepath)
        all_results.extend(results)

        for r in results:
            if r['判定'] in ('FAIL', 'WARN'):
                has_failure = True

        # 振り直し案を生成
        try:
            wb = load_workbook(filepath, data_only=True)
            event_nos = []
            for sheet in wb.worksheets:
                if 'イベント記述書' in norm(sheet.title):
                    event_nos.extend(extract_event_nos_from_event_sheet(sheet))
            wb.close()
            plan = generate_renumbering_plan(event_nos)
            if plan:
                renumbering_plans[filepath.name] = plan
        except Exception:
            pass

    output_path = Path(args.output)
    write_report(all_results, renumbering_plans, output_path)

    # サマリー表示
    pass_count = sum(1 for r in all_results if r['判定'] == 'PASS')
    warn_count = sum(1 for r in all_results if r['判定'] == 'WARN')
    fail_count = sum(1 for r in all_results if r['判定'] == 'FAIL')
    print(f'\n結果: PASS={pass_count}, WARN={warn_count}, FAIL={fail_count}')

    sys.exit(1 if has_failure else 0)


if __name__ == '__main__':
    main()

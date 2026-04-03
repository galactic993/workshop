#!/usr/bin/env python3
"""
バリデーション整合性チェッカー

画面設計書の項目記述書の制御ルールとメッセージ一覧の整合性をチェックする。

Usage:
    python check_validation_consistency.py <設計書ディレクトリ> [--output result.xlsx]
"""

import argparse
import re
import sys
from pathlib import Path
from typing import Dict, List

import unicodedata

import openpyxl
from openpyxl import load_workbook


def norm(s): return unicodedata.normalize('NFC', str(s))


# 制御ルールのキーワードとそれが必要とするメッセージのパターン
VALIDATION_RULES = [
    {
        'keyword': '必須',
        'rule_type': '必須チェック',
        'required_msg_pattern': re.compile(r'(必須|入力.+ください|入力内容に誤り)'),
        'description': '必須項目のエラーメッセージ'
    },
    {
        'keyword': '桁数',
        'rule_type': '桁数チェック',
        'required_msg_pattern': re.compile(r'(桁数|文字数|以内|入力内容に誤り)'),
        'description': '桁数制限のエラーメッセージ'
    },
    {
        'keyword': '半角数字',
        'rule_type': '形式チェック（半角数字）',
        'required_msg_pattern': re.compile(r'(半角数字|数字|形式|入力内容に誤り)'),
        'description': '半角数字形式のエラーメッセージ'
    },
    {
        'keyword': '半角英数字',
        'rule_type': '形式チェック（半角英数字）',
        'required_msg_pattern': re.compile(r'(半角英数字|英数字|形式|入力内容に誤り)'),
        'description': '半角英数字形式のエラーメッセージ'
    },
    {
        'keyword': '日付',
        'rule_type': '形式チェック（日付）',
        'required_msg_pattern': re.compile(r'(日付|形式|入力内容に誤り)'),
        'description': '日付形式のエラーメッセージ'
    },
]


def get_cell_value(sheet, row: int, col: int) -> str:
    value = sheet.cell(row=row, column=col).value
    return str(value).strip() if value is not None else ""


def extract_items_with_controls(sheet) -> List[Dict]:
    """項目記述書から制御内容を持つ項目を抽出"""
    items = []

    no_header = get_cell_value(sheet, 4, 2)
    if not no_header or 'No' not in no_header:
        return items

    row = 5
    empty_count = 0
    while row <= sheet.max_row:
        no = get_cell_value(sheet, row, 2)
        if not no:
            empty_count += 1
            row += 1
            if empty_count >= 20:
                break
            continue
        empty_count = 0

        control_detail = get_cell_value(sheet, row, 35)  # col35: 制御内容
        if control_detail:
            items.append({
                'row': row,
                'no': no,
                'name': get_cell_value(sheet, row, 7),
                'control_detail': control_detail,
            })
        row += 1

    return items


def extract_messages(sheet) -> List[str]:
    """メッセージ一覧からメッセージテキストを抽出"""
    messages = []

    header_row = None
    for row in range(1, min(15, sheet.max_row + 1)):
        if get_cell_value(sheet, row, 2) in ('No.', 'No'):
            header_row = row
            break

    if header_row is None:
        return messages

    row = header_row + 1
    empty_count = 0
    while row <= sheet.max_row:
        no = get_cell_value(sheet, row, 2)
        if not no:
            empty_count += 1
            row += 1
            if empty_count >= 20:
                break
            continue
        empty_count = 0

        msg = get_cell_value(sheet, row, 30)  # col30: メッセージ
        if msg:
            messages.append(msg)
        row += 1

    return messages


def check_file(filepath: Path) -> List[Dict]:
    """1つの画面設計書ファイルをチェックする"""
    results = []
    filename = filepath.name

    try:
        wb = load_workbook(filepath, data_only=True)
    except Exception as e:
        return [{'ファイル名': filename, '項目名': '-', '制御ルール': '-',
                 '行番号': '-', '詳細': f'読み込みエラー: {e}', '判定': 'FAIL'}]

    items_with_controls = []
    all_messages = []

    for sheet in wb.worksheets:
        if norm(sheet.title) == '項目記述書':
            items_with_controls.extend(extract_items_with_controls(sheet))
        elif 'メッセージ一覧' in norm(sheet.title):
            all_messages.extend(extract_messages(sheet))

    wb.close()

    if not items_with_controls:
        return [{'ファイル名': filename, '項目名': '-', '制御ルール': '-',
                 '行番号': '-', '詳細': '制御内容を持つ項目が見つかりません', '判定': 'PASS'}]

    # 各制御ルールキーワードについてメッセージ存在確認
    # 「入力内容に誤りがあります」という汎用メッセージがあれば多くのケースをカバー
    has_generic_input_error = any(
        '入力内容に誤りがあります' in m for m in all_messages
    )

    checked_rule_types = set()
    for item in items_with_controls:
        control = item['control_detail']
        for rule in VALIDATION_RULES:
            if rule['keyword'] not in control:
                continue
            if rule['rule_type'] in checked_rule_types:
                continue
            checked_rule_types.add(rule['rule_type'])

            # 対応するメッセージを探す
            found = any(rule['required_msg_pattern'].search(m) for m in all_messages)
            if found or has_generic_input_error:
                results.append({
                    'ファイル名': filename,
                    '項目名': item['name'],
                    '制御ルール': rule['rule_type'],
                    '行番号': item['row'],
                    '詳細': f'{rule["description"]}が存在します',
                    '判定': 'PASS'
                })
            else:
                results.append({
                    'ファイル名': filename,
                    '項目名': item['name'],
                    '制御ルール': rule['rule_type'],
                    '行番号': item['row'],
                    '詳細': f'{rule["description"]}がメッセージ一覧に見つかりません',
                    '判定': 'WARN'
                })

    if not results:
        results.append({
            'ファイル名': filename, '項目名': '-', '制御ルール': '-',
            '行番号': '-', '詳細': 'チェック対象の制御ルールが見つかりません', '判定': 'PASS'
        })

    return results


def write_report(all_results: List[Dict], output_path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'チェック結果'

    headers = ['ファイル名', '項目名', '制御ルール', '行番号', '詳細', '判定']
    ws.append(headers)
    for row in all_results:
        ws.append([row.get(h, '') for h in headers])

    wb.save(output_path)
    print(f'レポート出力: {output_path}')


def main():
    parser = argparse.ArgumentParser(description='バリデーション整合性チェッカー')
    parser.add_argument('design_dir', help='設計書ディレクトリのパス')
    parser.add_argument('--output', default='check_validation_consistency_result.xlsx')
    args = parser.parse_args()

    design_dir = Path(args.design_dir)
    if not design_dir.is_dir():
        print(f'エラー: {design_dir}', file=sys.stderr)
        sys.exit(1)

    screen_files = [f for f in design_dir.glob('*.xlsx')
                    if '画面設計書' in norm(f.name) and not f.name.startswith('~$')]

    if not screen_files:
        print('画面設計書が見つかりませんでした。', file=sys.stderr)
        sys.exit(1)

    print(f'対象ファイル数: {len(screen_files)}')

    all_results = []
    has_failure = False

    for filepath in sorted(screen_files):
        print(f'チェック中: {filepath.name}')
        results = check_file(filepath)
        all_results.extend(results)
        for r in results:
            if r['判定'] in ('FAIL', 'WARN'):
                has_failure = True

    write_report(all_results, Path(args.output))

    pass_count = sum(1 for r in all_results if r['判定'] == 'PASS')
    warn_count = sum(1 for r in all_results if r['判定'] == 'WARN')
    fail_count = sum(1 for r in all_results if r['判定'] == 'FAIL')
    print(f'\n結果: PASS={pass_count}, WARN={warn_count}, FAIL={fail_count}')

    sys.exit(1 if has_failure else 0)


if __name__ == '__main__':
    main()

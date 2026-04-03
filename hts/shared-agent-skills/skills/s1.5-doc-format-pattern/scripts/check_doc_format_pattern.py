#!/usr/bin/env python3
"""
設計書フォーマットパターン検出チェッカー

複数の画面設計書を横断分析し、シート構成・ヘッダー位置・採番パターン・
メッセージパターンの共通性と差異を検出する。

Usage:
    python check_doc_format_pattern.py <設計書ディレクトリ> [--output result.xlsx]
"""

import argparse
import re
import sys
from collections import Counter
from pathlib import Path
from typing import Dict, List, Tuple

import unicodedata

import openpyxl
from openpyxl import load_workbook


def norm(s): return unicodedata.normalize('NFC', str(s))


# 期待するシート名（部分一致）
EXPECTED_SHEETS = [
    '表紙', '変更履歴', '画面概要', 'イベント記述書', '項目記述書',
    '参照仕様', 'DB登録値', '表示切替項目', 'メッセージ一覧'
]


def get_cell_value(sheet, row: int, col: int) -> str:
    value = sheet.cell(row=row, column=col).value
    return str(value).strip() if value is not None else ""


def analyze_sheet_structure(filepath: Path) -> Dict:
    """ファイルのシート構成を分析する"""
    try:
        wb = load_workbook(filepath, data_only=True, read_only=True)
    except Exception as e:
        return {'error': str(e)}

    sheet_names = wb.sheetnames
    wb.close()

    # 期待シートの存在確認
    missing = []
    for expected in EXPECTED_SHEETS:
        found = any(expected in name for name in sheet_names)
        if not found:
            missing.append(expected)

    return {
        'sheets': sheet_names,
        'sheet_count': len(sheet_names),
        'missing_sheets': missing,
    }


def analyze_event_numbering(filepath: Path) -> Dict:
    """EVENT No採番パターンを分析する"""
    try:
        wb = load_workbook(filepath, data_only=True)
    except Exception as e:
        return {'error': str(e)}

    event_nos = []
    pattern = re.compile(r'^EVENT(\d{4})$')

    for sheet in wb.worksheets:
        if 'イベント記述書' not in norm(sheet.title):
            continue
        for row in range(6, sheet.max_row + 1):
            val = get_cell_value(sheet, row, 2)
            m = pattern.match(val)
            if m:
                event_nos.append(int(m.group(1)))

    wb.close()

    if not event_nos:
        return {'count': 0, 'min': None, 'max': None, 'format': 'EVENT\\d{4}'}

    return {
        'count': len(event_nos),
        'min': min(event_nos),
        'max': max(event_nos),
        'format': 'EVENT\\d{4}',
    }


def analyze_messages(filepath: Path) -> List[str]:
    """メッセージ一覧からメッセージを抽出する"""
    try:
        wb = load_workbook(filepath, data_only=True)
    except Exception:
        return []

    messages = []
    for sheet in wb.worksheets:
        if 'メッセージ一覧' not in norm(sheet.title):
            continue
        header_row = None
        for row in range(1, min(15, sheet.max_row + 1)):
            if get_cell_value(sheet, row, 2) in ('No.', 'No'):
                header_row = row
                break
        if header_row is None:
            continue
        row = header_row + 1
        empty_count = 0
        while row <= sheet.max_row:
            no = get_cell_value(sheet, row, 2)
            if not no:
                empty_count += 1
                row += 1
                if empty_count >= 20:
                    break
                continue
            empty_count = 0
            msg = get_cell_value(sheet, row, 30)
            if msg:
                messages.append(msg)
            row += 1

    wb.close()
    return messages


def classify_message_pattern(msg: str) -> str:
    """メッセージをパターン分類する"""
    if re.search(r'権限がありません', msg):
        return 'MSG-AUTH'
    if re.search(r'の(取得|検索|読込)に失敗しました。時間を空けて', msg):
        return 'MSG-FETCH-FAIL'
    if re.search(r'に失敗しました$', msg):
        return 'MSG-PROC-FAIL'
    if re.search(r'^入力内容に誤りがあります', msg):
        return 'MSG-INPUT-ERR'
    if re.search(r'^該当する.+が見つかりません', msg):
        return 'MSG-NOT-FOUND'
    if re.search(r'しました$', msg):
        return 'MSG-SUCCESS'
    return 'MSG-UNKNOWN'


def write_report(
    file_analyses: List[Dict],
    output_path: Path
) -> None:
    wb = openpyxl.Workbook()

    # シート1: シート構成比較
    ws1 = wb.active
    ws1.title = 'シート構成比較'
    ws1.append(['ファイル名', 'シート数', '欠落シート', '判定'])
    for fa in file_analyses:
        name = fa['filename']
        struct = fa['structure']
        if 'error' in struct:
            ws1.append([name, '-', struct['error'], 'FAIL'])
            continue
        missing = ', '.join(struct['missing_sheets']) if struct['missing_sheets'] else 'なし'
        judgment = 'WARN' if struct['missing_sheets'] else 'PASS'
        ws1.append([name, struct['sheet_count'], missing, judgment])

    # シート2: EVENT No採番分析
    ws2 = wb.create_sheet('EVENT No採番分析')
    ws2.append(['ファイル名', 'EVENT No件数', '最小番号', '最大番号', 'フォーマット', '判定'])
    for fa in file_analyses:
        name = fa['filename']
        ev = fa['event_numbering']
        if 'error' in ev:
            ws2.append([name, '-', '-', '-', '-', 'FAIL'])
            continue
        judgment = 'PASS' if ev['count'] > 0 else 'WARN'
        ws2.append([name, ev['count'],
                    ev['min'] if ev['min'] is not None else '-',
                    ev['max'] if ev['max'] is not None else '-',
                    ev['format'], judgment])

    # シート3: メッセージパターン頻度
    ws3 = wb.create_sheet('メッセージパターン頻度')
    all_patterns: Counter = Counter()
    for fa in file_analyses:
        for msg in fa['messages']:
            pat = classify_message_pattern(msg)
            all_patterns[pat] += 1

    ws3.append(['パターンID', '件数'])
    for pat, count in all_patterns.most_common():
        ws3.append([pat, count])

    # シート4: 全メッセージ一覧（ファイル別）
    ws4 = wb.create_sheet('メッセージ全件')
    ws4.append(['ファイル名', 'メッセージ', 'パターンID'])
    for fa in file_analyses:
        for msg in fa['messages']:
            ws4.append([fa['filename'], msg, classify_message_pattern(msg)])

    wb.save(output_path)
    print(f'レポート出力: {output_path}')


def main():
    parser = argparse.ArgumentParser(description='設計書フォーマットパターン検出チェッカー')
    parser.add_argument('design_dir', help='設計書ディレクトリのパス')
    parser.add_argument('--output', default='check_doc_format_pattern_result.xlsx')
    args = parser.parse_args()

    design_dir = Path(args.design_dir)
    if not design_dir.is_dir():
        print(f'エラー: {design_dir}', file=sys.stderr)
        sys.exit(1)

    screen_files = [f for f in design_dir.glob('*.xlsx')
                    if '画面設計書' in norm(f.name) and not f.name.startswith('~$')]

    if not screen_files:
        print('画面設計書が見つかりませんでした。', file=sys.stderr)
        sys.exit(1)

    print(f'対象ファイル数: {len(screen_files)}')

    file_analyses = []
    has_warning = False

    for filepath in sorted(screen_files):
        print(f'分析中: {filepath.name}')
        struct = analyze_sheet_structure(filepath)
        ev = analyze_event_numbering(filepath)
        messages = analyze_messages(filepath)

        file_analyses.append({
            'filename': filepath.name,
            'structure': struct,
            'event_numbering': ev,
            'messages': messages,
        })

        if struct.get('missing_sheets'):
            has_warning = True

    write_report(file_analyses, Path(args.output))

    print(f'\n分析完了: {len(file_analyses)}ファイル')
    sys.exit(1 if has_warning else 0)


if __name__ == '__main__':
    main()

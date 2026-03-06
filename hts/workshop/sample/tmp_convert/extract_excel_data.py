#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import openpyxl
from openpyxl.utils import get_column_letter
import json

excel_file = "/Users/izutanikazuki/kzp/fileMaker/training/sample/編-テーブル定義書_1012.xlsx"

# Excelファイルを読み込む
print("=== Excelファイル読み込み開始 ===\n")
wb = openpyxl.load_workbook(excel_file, data_only=False)

# すべてのシート名を取得
sheet_names = wb.sheetnames
print(f"【シート一覧】\n合計シート数: {len(sheet_names)}\n")
for i, name in enumerate(sheet_names, 1):
    print(f"  {i}. {name}")

print("\n" + "="*80 + "\n")

# 各シートのデータを抽出
for sheet_idx, sheet_name in enumerate(sheet_names, 1):
    ws = wb[sheet_name]
    print(f"【シート #{sheet_idx}: {sheet_name}】")
    print(f"寸法: {ws.dimensions}")
    print(f"最大行: {ws.max_row}, 最大列: {ws.max_column}")
    print()

    # マージされたセルの情報を取得
    if ws.merged_cells:
        print(f"マージされたセル ({len(ws.merged_cells.ranges)} 個):")
        for merged_range in ws.merged_cells.ranges:
            print(f"  {merged_range}")
        print()

    # すべてのデータを表形式で表示
    print("【データ内容】")
    print()

    # ヘッダー行（列番号）を表示
    col_headers = []
    for col in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col)
        col_headers.append(col_letter)

    header_str = "行番号 | " + " | ".join(f"{ch:15}" for ch in col_headers)
    print(header_str)
    print("-" * len(header_str))

    # データ行を表示
    for row in range(1, ws.max_row + 1):
        row_data = [f"{row:4}"]
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            value = cell.value

            # セルの値をフォーマット
            if value is None:
                cell_display = "(空)"
            elif isinstance(value, str):
                cell_display = value[:20] if len(value) > 20 else value
            else:
                cell_display = str(value)[:20]

            row_data.append(f"{cell_display:15}")

        print(" | ".join(row_data))

    print()
    print("="*80)
    print()

# 詳細な JSON 形式でのエクスポート
print("\n【詳細データエクスポート（JSON形式）】\n")

output_data = {}

for sheet_name in sheet_names:
    ws = wb[sheet_name]
    sheet_data = {
        "dimensions": str(ws.dimensions),
        "max_row": ws.max_row,
        "max_column": ws.max_column,
        "merged_cells": list(str(r) for r in ws.merged_cells.ranges),
        "rows": []
    }

    for row in range(1, ws.max_row + 1):
        row_values = []
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell_info = {
                "column": get_column_letter(col),
                "column_index": col,
                "value": cell.value,
                "data_type": cell.data_type,
            }
            row_values.append(cell_info)
        sheet_data["rows"].append(row_values)

    output_data[sheet_name] = sheet_data

# JSON形式で保存
json_output_path = "/Users/izutanikazuki/kzp/fileMaker/training/sample/tmp_convert/excel_data.json"
with open(json_output_path, 'w', encoding='utf-8') as f:
    json.dump(output_data, f, ensure_ascii=False, indent=2)

print(f"JSON形式のデータを保存しました: {json_output_path}\n")

# CSV形式での出力（最初のシートのみ）
if sheet_names:
    first_sheet = sheet_names[0]
    ws = wb[first_sheet]

    csv_output_path = "/Users/izutanikazuki/kzp/fileMaker/training/sample/tmp_convert/excel_data_first_sheet.csv"
    with open(csv_output_path, 'w', encoding='utf-8-sig') as f:
        for row in range(1, ws.max_row + 1):
            row_values = []
            for col in range(1, ws.max_column + 1):
                value = ws.cell(row=row, column=col).value
                if value is None:
                    row_values.append("")
                else:
                    row_values.append(str(value).replace(',', '，').replace('\n', ' '))
            f.write(",".join(row_values) + "\n")

    print(f"最初のシート ({first_sheet}) をCSV形式で保存しました: {csv_output_path}\n")

print("=== 処理完了 ===")
wb.close()

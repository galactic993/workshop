#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import openpyxl
from openpyxl.utils import get_column_letter
import pickle

excel_file = "/Users/izutanikazuki/kzp/fileMaker/training/sample/編-テーブル定義書_1012.xlsx"

# Excelファイルを読み込む
print("=== Excelファイル読み込み開始 ===\n")
wb = openpyxl.load_workbook(excel_file, data_only=False)

# すべてのシート名を取得
sheet_names = wb.sheetnames
print(f"【シート一覧】\n合計シート数: {len(sheet_names)}\n")
for i, name in enumerate(sheet_names, 1):
    print(f"  {i}. {name}")

print("\n" + "="*80 + "\n")

# 各シートの基本情報を表示
for sheet_idx, sheet_name in enumerate(sheet_names, 1):
    ws = wb[sheet_name]
    print(f"【シート #{sheet_idx}: {sheet_name}】")
    print(f"  寸法: {ws.dimensions}")
    print(f"  最大行: {ws.max_row}, 最大列: {ws.max_column}")

    # マージセルの情報
    if ws.merged_cells:
        print(f"  マージセル数: {len(ws.merged_cells.ranges)}")

    # データを抽出（先頭20行のサンプル）
    print(f"  データサンプル（先頭20行）:")
    for row in range(1, min(21, ws.max_row + 1)):
        row_data = []
        for col in range(1, min(11, ws.max_column + 1)):  # 最初の10列
            cell = ws.cell(row=row, column=col)
            value = cell.value
            if value is None:
                row_data.append("")
            else:
                val_str = str(value)[:25]
                row_data.append(val_str)
        print(f"    行{row}: {row_data}")

    print()

# シート別に詳細な情報をテキストファイルで保存
print("\n【詳細データをテキストファイルで保存中...】\n")

for sheet_name in sheet_names:
    ws = wb[sheet_name]
    output_path = f"/Users/izutanikazuki/kzp/fileMaker/training/sample/tmp_convert/sheet_{sheet_name}.txt"

    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(f"=== シート: {sheet_name} ===\n\n")
        f.write(f"寸法: {ws.dimensions}\n")
        f.write(f"最大行: {ws.max_row}, 最大列: {ws.max_column}\n\n")

        if ws.merged_cells:
            f.write(f"マージセル ({len(ws.merged_cells.ranges)} 個):\n")
            for merged_range in ws.merged_cells.ranges:
                f.write(f"  {merged_range}\n")
            f.write("\n")

        f.write("【データ内容】\n\n")

        # ヘッダー行（列番号）
        header = "行番号"
        for col in range(1, ws.max_column + 1):
            header += f" | {get_column_letter(col)}"
        f.write(header + "\n")
        f.write("-" * len(header) + "\n\n")

        # データテーブル
        for row in range(1, ws.max_row + 1):
            row_values = [str(row)]
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                value = cell.value
                if value is None:
                    row_values.append("")
                else:
                    # 改行を削除
                    val_str = str(value).replace('\n', ' ')
                    row_values.append(val_str)

            # パイプ区切りで出力
            f.write(" | ".join(row_values) + "\n")

    print(f"  保存: {output_path}")

# サマリーレポートを作成
summary_path = "/Users/izutanikazuki/kzp/fileMaker/training/sample/tmp_convert/EXCEL_SUMMARY.md"
with open(summary_path, 'w', encoding='utf-8') as f:
    f.write("# Excel ファイル抽出レポート\n\n")
    f.write(f"ファイル: 編-テーブル定義書_1012.xlsx\n\n")

    f.write("## シート一覧\n\n")
    f.write("| No. | シート名 | 行数 | 列数 | マージセル数 |\n")
    f.write("|-----|---------|------|------|----------------|\n")

    for idx, sheet_name in enumerate(sheet_names, 1):
        ws = wb[sheet_name]
        merged_count = len(ws.merged_cells.ranges) if ws.merged_cells else 0
        f.write(f"| {idx} | {sheet_name} | {ws.max_row} | {ws.max_column} | {merged_count} |\n")

    f.write("\n## データ構造\n\n")
    f.write("各シートは以下の構造を持ちます：\n\n")

    # 最初のいくつかのシートの詳細情報を表示
    for sheet_name in sheet_names[:5]:
        ws = wb[sheet_name]
        f.write(f"\n### {sheet_name}\n\n")
        f.write(f"- 寸法: {ws.dimensions}\n")
        f.write(f"- 最大行: {ws.max_row}\n")
        f.write(f"- 最大列: {ws.max_column}\n")
        if ws.merged_cells:
            f.write(f"- マージセル: {len(ws.merged_cells.ranges)} 個\n")

        # サンプルデータ（最初の5行、最初の8列）
        f.write(f"- サンプルデータ (先頭5行x8列):\n\n")
        f.write("| 行 |")
        for col in range(1, 9):
            f.write(f" {get_column_letter(col)} |")
        f.write("\n")
        f.write("|----" + ("|----|" * 8) + "\n")

        for row in range(1, min(6, ws.max_row + 1)):
            f.write(f"| {row} |")
            for col in range(1, 9):
                cell = ws.cell(row=row, column=col)
                value = cell.value
                if value is None:
                    f.write(" |")
                else:
                    val_str = str(value)[:20].replace('\n', ' ')
                    f.write(f" {val_str} |")
            f.write("\n")

print(f"\n  保存: {summary_path}")

# Pickleフォーマットでもデータをバックアップ保存（JSONの代わり）
print("\n【データをPickle形式でバックアップ...】\n")

backup_data = {}
for sheet_name in sheet_names:
    ws = wb[sheet_name]
    sheet_data = {
        "dimensions": str(ws.dimensions),
        "max_row": ws.max_row,
        "max_column": ws.max_column,
        "merged_cells": list(str(r) for r in ws.merged_cells.ranges) if ws.merged_cells else [],
        "data": []
    }

    for row in range(1, ws.max_row + 1):
        row_values = []
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            # 値をそのまま保存（型情報も保持）
            row_values.append(cell.value)
        sheet_data["data"].append(row_values)

    backup_data[sheet_name] = sheet_data

pickle_path = "/Users/izutanikazuki/kzp/fileMaker/training/sample/tmp_convert/excel_complete_data.pkl"
with open(pickle_path, 'wb') as f:
    pickle.dump(backup_data, f)

print(f"  保存: {pickle_path}")

print("\n=== 処理完了 ===")
print(f"\n出力ファイル:")
print(f"  1. sheet_*.txt (各シートを詳細データで保存)")
print(f"  2. {summary_path} (概要レポート)")
print(f"  3. {pickle_path} (Pickle形式バックアップ)")

wb.close()

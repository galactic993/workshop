#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import openpyxl
from openpyxl.utils import get_column_letter
import json

excel_file = "/Users/izutanikazuki/kzp/fileMaker/training/sample/編-テーブル定義書_1012.xlsx"

# Excelファイルを読み込む
print("=== Excelファイル読み込み開始 ===\n")
wb = openpyxl.load_workbook(excel_file, data_only=False)

# すべてのシート名を取得
sheet_names = wb.sheetnames
print(f"【シート一覧】\n合計シート数: {len(sheet_names)}\n")
for i, name in enumerate(sheet_names, 1):
    print(f"  {i}. {name}")

print("\n" + "="*80 + "\n")

# 各シートの基本情報を表示
for sheet_idx, sheet_name in enumerate(sheet_names, 1):
    ws = wb[sheet_name]
    print(f"【シート #{sheet_idx}: {sheet_name}】")
    print(f"  寸法: {ws.dimensions}")
    print(f"  最大行: {ws.max_row}, 最大列: {ws.max_column}")

    # マージセルの情報
    if ws.merged_cells:
        print(f"  マージセル数: {len(ws.merged_cells.ranges)}")

    # データを抽出（先頭20行のサンプル）
    print(f"  データサンプル（先頭20行）:")
    for row in range(1, min(21, ws.max_row + 1)):
        row_data = []
        for col in range(1, min(11, ws.max_column + 1)):  # 最初の10列
            cell = ws.cell(row=row, column=col)
            value = cell.value
            if value is None:
                row_data.append("")
            else:
                val_str = str(value)[:25]
                row_data.append(val_str)
        print(f"    行{row}: {row_data}")

    print()

# 詳細データを JSON で保存
print("\n【詳細データをJSON形式で保存中...】\n")

output_data = {}

for sheet_name in sheet_names:
    ws = wb[sheet_name]
    sheet_data = {
        "dimensions": str(ws.dimensions),
        "max_row": ws.max_row,
        "max_column": ws.max_column,
        "merged_cells": list(str(r) for r in ws.merged_cells.ranges) if ws.merged_cells else [],
        "data": []
    }

    for row in range(1, ws.max_row + 1):
        row_values = []
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            row_values.append(cell.value)
        sheet_data["data"].append(row_values)

    output_data[sheet_name] = sheet_data

# JSON保存
json_output_path = "/Users/izutanikazuki/kzp/fileMaker/training/sample/tmp_convert/excel_complete_data.json"
with open(json_output_path, 'w', encoding='utf-8') as f:
    json.dump(output_data, f, ensure_ascii=False, indent=2)

print(f"完全なデータを JSON で保存: {json_output_path}\n")

# シート別に詳細な情報を表示・保存
for sheet_name in sheet_names:
    ws = wb[sheet_name]
    output_path = f"/Users/izutanikazuki/kzp/fileMaker/training/sample/tmp_convert/sheet_{sheet_name}.txt"

    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(f"=== シート: {sheet_name} ===\n\n")
        f.write(f"寸法: {ws.dimensions}\n")
        f.write(f"最大行: {ws.max_row}, 最大列: {ws.max_column}\n\n")

        if ws.merged_cells:
            f.write(f"マージセル ({len(ws.merged_cells.ranges)} 個):\n")
            for merged_range in ws.merged_cells.ranges:
                f.write(f"  {merged_range}\n")
            f.write("\n")

        f.write("【データ内容】\n\n")

        # データテーブル
        for row in range(1, ws.max_row + 1):
            row_values = []
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                value = cell.value
                if value is None:
                    row_values.append("")
                else:
                    row_values.append(str(value))

            # CSV形式で出力
            f.write(",".join(row_values) + "\n")

    print(f"  保存: {output_path}")

print("\n=== 処理完了 ===")
print(f"\n出力ファイル:")
print(f"  1. {json_output_path} (全データをJSON形式で保存)")
print(f"  2. sheet_*.txt (各シートをCSV形式で保存)")

wb.close()
